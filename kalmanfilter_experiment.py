import openpyxl
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

# Path file input dan output
input_file = r'files\DOSS.xlsx'
output_file = r'files\hasil_kalman.xlsx'

# Load data dengan openpyxl
workbook = openpyxl.load_workbook(input_file)
worksheet = workbook.active

# Ambil data kolom A (DOSS1), B (DOSS2), dan C (DOSS3)
rss_doss1, rss_doss2, rss_doss3 = [], [], []
for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
    rss_doss1.append(int(row[0].value) if row[0].value else 0)
    rss_doss2.append(int(row[1].value) if row[1].value else 0)
    rss_doss3.append(int(row[2].value) if row[2].value else 0)

# Implementasi Kalman Filter dengan adaptasi parameter dan inisialisasi

def adaptive_kalman_filter(data):
    n_iter = len(data)
    xhat = np.zeros(n_iter)
    P = np.zeros(n_iter)
    K = np.zeros(n_iter)

    # Inisialisasi dinamis
    xhat[0] = data[0]
    P[0] = np.var(data)  

    # Adaptasi parameter secara dinamis
    Q = np.var(data) * 0.01
    R = np.var(data) * 0.1

    for k in range(1, n_iter):
        # Time Update
        xhatminus = xhat[k-1]
        Pminus = P[k-1] + Q

        # Measurement Update
        K[k] = Pminus / (Pminus + R)
        xhat[k] = xhatminus + K[k] * (data[k] - xhatminus)
        P[k] = (1 - K[k]) * Pminus

    return xhat

# Apply adaptive Kalman filter
kalman_doss1 = adaptive_kalman_filter(rss_doss1)
kalman_doss2 = adaptive_kalman_filter(rss_doss2)
kalman_doss3 = adaptive_kalman_filter(rss_doss3)

# Simpan hasil Kalman Filter ke Excel
df_output = pd.DataFrame({
    'Kalman DOSS1': kalman_doss1,
    'Kalman DOSS2': kalman_doss2,
    'Kalman DOSS3': kalman_doss3
})
df_output.to_excel(output_file, index=False, engine='openpyxl')

# Hitung MSE (Mean Squared Error)
mse_doss1 = np.mean((np.array(rss_doss1) - kalman_doss1)**2)
mse_doss2 = np.mean((np.array(rss_doss2) - kalman_doss2)**2)
mse_doss3 = np.mean((np.array(rss_doss3) - kalman_doss3)**2)

print(f"Adaptive MSE DOSS1: {mse_doss1:.4f}")
print(f"Adaptive MSE DOSS2: {mse_doss2:.4f}")
print(f"Adaptive MSE DOSS3: {mse_doss3:.4f}")

# Fungsi plot hasil Kalman Filter
#def plot_kalman_results(original, kalman_filtered, label):
#    plt.figure(figsize=(12, 6))
#    plt.plot(original, label=f'Original {label}', linestyle='--', alpha=0.7)
#    plt.plot(kalman_filtered, label=f'Kalman Filtered {label}', linewidth=2)
#    plt.xlabel('Data Index')
#    plt.ylabel('DOSS (dBm)')
#    plt.legend()
#    plt.title(f'Kalman Filter pada {label}')
#    plt.grid(True)
#    plt.show()

# Plot hasil Kalman Filter
#plot_kalman_results(rss_doss1, kalman_doss1, 'DOSS1')
#plot_kalman_results(rss_doss2, kalman_doss2, 'DOSS2')
#plot_kalman_results(rss_doss3, kalman_doss3, 'DOSS3')

print(f"Hasil Adaptive Kalman Filter disimpan pada: {output_file}")
