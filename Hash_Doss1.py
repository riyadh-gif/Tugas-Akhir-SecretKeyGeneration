import time
import pandas as pd
import hashlib
import pyaes
import binascii
import os
import csv
import numpy as np
from openpyxl import load_workbook, Workbook  # openpyxl for .xlsx files
import xlrd  # xlrd for .xls files
# =========================================================================
# ============================== UNIVERSAL HASH ===========================

start5 = time.time()

# Using xlrd for .xls and openpyxl for .xlsx
file_path1 = "E:\\PENS\\Semester 7\\Progress PA\\codingan\\Program\\Hasil_BCH_DOSS1.xls"
file_path2 = "E:\\PENS\\Semester 7\\Progress PA\\codingan\\Program\\Hasil_BCH_Bob_doss2.xls"

try:
    # Check if it's an .xls file and open it using xlrd
    if file_path1.endswith('.xls'):
        workbook = xlrd.open_workbook(file_path1)  # Read .xls file using xlrd
        sheet_names = workbook.sheet_names()  # Get all sheet names to check the correct one
        print(f"Available sheet names in .xls file: {sheet_names}")
        worksheet = workbook.sheet_by_name('Sheet1')  # Change to 'Sheet1' as per your file
    elif file_path1.endswith('.xlsx'):
        workbook = load_workbook(file_path1, data_only=True)  # Read .xlsx file using openpyxl
        worksheet = workbook['HasilBCHalice']  # Change to the correct sheet name
    
    # Check if the second file exists and is the correct format
    if not os.path.exists(file_path2):
        print(f"File {file_path2} not found. Please ensure the file exists.")
        exit(1)
    
    # Read second file (assuming it's .xlsx)
    if file_path2.endswith('.xls'):
        workbook1 = xlrd.open_workbook(file_path2)  # Read .xls file using xlrd
        worksheet1 = workbook1.sheet_by_name('HasilBCHbob')  # Change to the correct sheet name
    elif file_path2.endswith('.xlsx'):
        workbook1 = load_workbook(file_path2, data_only=True)  # Read .xlsx file using openpyxl
        worksheet1 = workbook1['HasilBCHbob']  # Change to the correct sheet name

except FileNotFoundError as e:
    print(f"Excel file not found: {e}")
    exit(1)
except KeyError as e:
    print(f"Error: Sheet not found. Please check the sheet name. {e}")
    exit(1)
except Exception as e:
    print(f"Error reading Excel files: {e}")
    exit(1)

# Extract data from worksheet for .xls (using xlrd method)
aliice = []
if isinstance(workbook, xlrd.Book):
    for row_num in range(1, worksheet.nrows):  # Skip header row
        aliice.append(worksheet.cell_value(row_num, 0))  # Extract the first column
else:
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=1):
        aliice.append(row[0].value)

print(f"Panjang Input UnivHASH ALICE {len(aliice)}")

# Continue with the rest of your code...


# Continue with the rest of your code...

# Continue with the rest of your code...


# Calculate the number of keys needed
jumlahkeya = len(aliice) // 128
jumlahkey = jumlahkeya
ukuranhash = 128
aaaa = len(aliice) % 128
lenalice = len(aliice) - aaaa

# Trim the data if necessary
for i in range(0, aaaa):
    del aliice[lenalice]

# Read the hash table from a CSV file
Hashtab = []
with open('Hashtable128.csv', newline='') as f:
    reader = csv.reader(f)
    for row in reader:
        Hashtab.append(row)
Hashtable = np.array(Hashtab)

# Generate the keys
key1 = []
for i in range(jumlahkey):
    mat1 = []
    aaa = aliice[(i * ukuranhash): (ukuranhash * (i + 1))]
    print(f"Key-{i + 1}: Panjang data adalah {len(aaa)} dan ukuran HashTable yaitu {ukuranhash} x {len(aaa)}")
    for x in range(0, len(Hashtable)):
        row = []
        total = 0
        for y in range(0, len(aaa)):
            total = total + (int(Hashtable[x][y]) * aaa[y])
            row = total % 2
        mat1.append(int(row))
    key1.append(mat1)
    print(f"Jumlah KEY ALICE Sekarang: {len(key1)}")

# Flatten the keys into a single array
v = 0
ax = [0] * 128 * jumlahkey
for i in range(jumlahkey):
    for j in range(128):
        ax[v] = key1[i][j]
        v = v + 1
univ = []
for i in range(0, len(ax)):
    univ.append(ax[i])

univ1 = np.array(univ)
univ2 = univ1.reshape(len(ax), 1)

# Save the results to a CSV file
with open("E:\\PENS\\Semester 7\\Progress PA\\codingan\\Program\\univhash_Alice_doss1.csv", "w", newline="") as fp:
    writer = csv.writer(fp, delimiter=",")
    writer.writerows(univ2)

# Save the results to an Excel file using openpyxl
wb = Workbook()  # Use Workbook() to create a new workbook
ws = wb.active
ws.title = "UnivHASH"
ws.cell(row=1, column=1, value="Alice")
for i in range(1, len(ax) + 1):
    ws.cell(row=i + 1, column=1, value=int(ax[i - 1]))
wb.save("Universal_Hash_Alice_dos1.xlsx")  # Save as .xlsx

end5 = time.time()
print(f'UNIVHASH Panjang bit hasil Universal Hash alice = {len(ax)}, bob= {len(ax)}')
print(f'UNIVHASH Jumlah hasil key yang dibangkitkan = {jumlahkey}')
print(f'Waktu Proses HASHING : {end5 - start5}')

# =========================================================================
# ============================== CEK NIST =================================
print('\n\n====================== CEK NIST ========================')
print('============================================================\n')

startnist = time.time()

# Read NIST test results from CSV
indeks = []
try:
    with open("E:\\PENS\\Semester 7\\Progress PA\\codingan\\Program\\output\\sudahujinist_Alice.csv", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            indeks.append(row)
except FileNotFoundError as e:
    print(f"NIST output file not found: {e}")
    exit(1)

prindek = [int(x[0]) + 1 for x in indeks]  # Assuming you want to increment the values in the index

endnist = time.time()
print('NIST Hasil prioritas index', prindek)
print(f'Waktu Proses Uji NIST : {endnist - startnist}')

# =========================================================================
# ================================= SHA-128 ===============================
print('\n\n====================== SHA-128 ========================')
print('============================================================\n')

start6 = time.time()

# SHA-128 Hashing
dataalice = []
hex1 = []
abc1 = []
# Assuming 'indeks' values need to be converted to integers
indeks = [int(x[0]) + 1 for x in indeks]  # Assuming x[0] is the value you want

# Check if 'indeks' is now an integer list, and use it correctly in the hash calculation
for j in range(len(indeks)):
    hash1 = [0] * 128
    for i in range(128):
        # Ensure indeks[j] is an integer before using it as an index
        hash1[i] = key1[0][int(indeks[j])]  # Convert to int just in case
    data1 = "".join(str(e) for e in hash1)
    someText1 = data1.encode("ascii")
    b1 = hashlib.sha1(someText1).hexdigest()
    hex1.append(b1)
    print(f'Hasil hash Alice Kunci-{indeks[j]} = {hex1[j]}')


# Save SHA128 results to an Excel file
wb = Workbook()  # Create a new workbook to store SHA-128 results
ws = wb.active
ws.title = "SHA128"
ws.cell(row=1, column=1, value="Alice")
for i in range(1, len(hex1) + 1):
    ws.cell(row=i + 1, column=1, value=hex1[i - 1])
wb.save('SHA128ALICE_I10_7030_97.xlsx')

end6 = time.time()
print(f'Waktu Proses HASHING : {end6 - start6}')

# =========================================================================
# ================================= AES-128 ===============================
print('\n\n===========~~~~~~~~~~~== AES ~~~~~~~~~~~~~~==============')
print('===========~~~~~~~~~~~=========~~~~~~~~~~~~~~==============\n')

start7 = time.time()

# AES encryption
for kuncinya in range(len(indeks)):
    if hex1[kuncinya] == hex1[kuncinya]:
        keybit = ''.join(str(e) for e in key1[indeks[kuncinya]])
        keyint = int(keybit, 2)
        hex_str = '%x' % keyint
        if len(hex_str) % 2 != 0:
            hex_str = '0' + hex_str
        keybyte = binascii.unhexlify(hex_str)
        print('Key Bob 1 (16 bytes) = ', keybyte)
        break

for kunci in range(len(indeks)):
    if hex1[kunci] == hex1[kunci]:
        keybitt = ''.join(str(e) for e in key1[indeks[kunci]])
        keyintt = int(keybitt, 2)
        hex_strt = '%x' % keyintt
        if len(hex_strt) % 2 != 0:
            hex_strt = '0' + hex_strt
        keybytee = binascii.unhexlify(hex_strt)
        print('Key Bob 2 (32 bytes) = ', keybytee)

        print('\n=================CTR=============')
        plaintext = "TES PESAN"
        aesctr = pyaes.AESModeOfOperationCTR(keybyte)
        aesctrr = pyaes.AESModeOfOperationCTR(keybytee)
        ciphertext = aesctr.encrypt(plaintext)
        ciphertextt = aesctrr.encrypt(plaintext)
        print('Ciphertext 1 key 1 = ', ciphertext)
        print('Ciphertext 2 key 2 = ', ciphertextt)
        break

end7 = time.time()
print(f'Waktu Proses Enkripsi : {end7 - start7}')
