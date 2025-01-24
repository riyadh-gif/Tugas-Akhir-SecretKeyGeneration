import openpyxl
import numpy as np
import time
from xlwt import Workbook
import matplotlib.pyplot as plt

start = time.time()

# =========================================================================
# ============================== KALMAN FILTER ============================
start1 = time.time()
interval = 0.11

# Menggunakan openpyxl untuk membaca file .xlsx
workbook = openpyxl.load_workbook(r'E:\PENS\Semester 7\Progress PA\codingan\Program\channelprob_skenario1.xlsx')
worksheet = workbook['Sheet1']  # Ganti dengan nama sheet jika berbeda

# Mengambil data dari kolom IP2, IP3, dan IP4
rss_ip2 = []
rss_ip3 = []
rss_ip4 = []

for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=3, values_only=True):
    # Validasi data kosong
    rss_ip2.append(int(row[0]) if row[0] is not None else 0)  # Kolom IP2
    rss_ip3.append(int(row[1]) if row[1] is not None else 0)  # Kolom IP3
    rss_ip4.append(int(row[2]) if row[2] is not None else 0)  # Kolom IP4

bb = 50  # Jumlah loop
total_data = len(rss_ip2)
aa = total_data // bb  # Jumlah blok data yang valid

# Potong data agar sesuai dengan reshape
rss_ip2 = rss_ip2[:aa * bb]
rss_ip3 = rss_ip3[:aa * bb]
rss_ip4 = rss_ip4[:aa * bb]

# Melakukan reshape data untuk IP2, IP3, IP4
ip2 = np.array(rss_ip2).reshape(aa, bb).T
ip3 = np.array(rss_ip3).reshape(aa, bb).T
ip4 = np.array(rss_ip4).reshape(aa, bb).T

a = 1     # a posteri error estimate
h = 1
R = 1     # Measurement Error Covariance Matrix
Q = 0.01  # Process Noise Covariance

xaposteriori_0 = -5  # Initial guesses
paposteriori_0 = 1

# Kalman Filter untuk IP2, IP3, IP4
def kalman_filter(rssival):
    xapriori = []
    residual = []
    papriori = []
    k = []
    paposteriori = []
    xaposteriori = []

    # Proses untuk baris pertama
    row1, row2, row3, row4, row5, row6 = [], [], [], [], [], []
    for m in range(0, aa):
        row1.append(a * xaposteriori_0)
        row2.append(rssival[0][m] - h * row1[m])
        row3.append(a * a * paposteriori_0 + Q)
        row4.append(row3[m] / (row3[m] + R))
        row5.append(row3[m] * (1 - row4[m]))
        row6.append(row1[m] + row4[m] * row2[m])
    xapriori.append(row1)
    residual.append(row2)
    papriori.append(row3)
    k.append(row4)
    paposteriori.append(row5)
    xaposteriori.append(row6)

    # Proses untuk baris-baris berikutnya
    for j in range(1, bb):
        row7, row8, row9, row10, row11, row12 = [], [], [], [], [], []
        for n in range(0, aa):
            row7.append(xaposteriori[j - 1][n])
            row8.append(rssival[j][n] - h * row7[n])
            row9.append(a * a * paposteriori[j - 1][n] + Q)
            row10.append(row9[n] / (row9[n] + R))
            row11.append(row9[n] * (1 - row10[n]))
            row12.append(row7[n] + row10[n] * row8[n])
        xapriori.append(row7)
        residual.append(row8)
        papriori.append(row9)
        k.append(row10)
        paposteriori.append(row11)
        xaposteriori.append(row12)
    
    return np.array(xaposteriori)

kalman_ip2 = kalman_filter(ip2).T.flatten()
kalman_ip3 = kalman_filter(ip3).T.flatten()
kalman_ip4 = kalman_filter(ip4).T.flatten()

# Menyimpan hasil ke file Excel
book = Workbook()
sheet1 = book.add_sheet('KalmanFilter')

sheet1.write(0, 0, 'IP2')
sheet1.write(0, 1, 'IP3')
sheet1.write(0, 2, 'IP4')

for i in range(len(kalman_ip2)):
    sheet1.write(i + 1, 0, int(kalman_ip2[i]))
    sheet1.write(i + 1, 1, int(kalman_ip3[i]))
    sheet1.write(i + 1, 2, int(kalman_ip4[i]))

book.save('Praproses_Kalman_IP2_IP3_IP4.xlsx')

end1 = time.time()
time_kalman = end1 - start1
print('Waktu komputasi kalman = {} seconds'.format(time_kalman))


##Nyoba ae


# Visualisasi hasil Kalman Filter
def plot_kalman_results(original, kalman_filtered, label):
    plt.figure(figsize=(10, 6))
    plt.plot(original, label=f'Original {label}', linestyle='--', alpha=0.7)
    plt.plot(kalman_filtered, label=f'Kalman Filtered {label}', linewidth=2)
    plt.title(f'Hasil Kalman Filter - {label}')
    plt.xlabel('Data Index')
    plt.ylabel('Value')
    plt.legend()
    plt.grid(True)
    plt.show()

# Plot untuk IP2
plot_kalman_results(rss_ip2, kalman_ip2, 'IP2')

# Plot untuk IP3
plot_kalman_results(rss_ip3, kalman_ip3, 'IP3')

# Plot untuk IP4
plot_kalman_results(rss_ip4, kalman_ip4, 'IP4')

