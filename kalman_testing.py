import openpyxl
import numpy as np
import time
import matplotlib.pyplot as plt

start = time.time()

# =========================================================================
# ============================== KALMAN FILTER ============================
start1 = time.time()
interval = 0.11

# Menggunakan openpyxl untuk membaca file .xlsx
workbook = openpyxl.load_workbook(r'E:\PENS\Semester 7\Progress PA\codingan\Program\channelprob_skenario1.xlsx')
worksheet = workbook['Sheet1']  # Ganti dengan nama sheet jika berbeda

# Mengambil data dari kolom IP1, IP2, IP3, dan IP4
rss_ip1 = []
rss_ip2 = []
rss_ip3 = []
rss_ip4 = []

for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=4, values_only=True):
    rss_ip1.append(int(row[0]) if row[0] is not None else 0)  # Kolom IP1
    rss_ip2.append(int(row[1]) if row[1] is not None else 0)  # Kolom IP2
    rss_ip3.append(int(row[2]) if row[2] is not None else 0)  # Kolom IP3
    rss_ip4.append(int(row[3]) if row[3] is not None else 0)  # Kolom IP4

bb = 50  # Jumlah loop
total_data = len(rss_ip1)
aa = total_data // bb  # Jumlah blok data yang valid

# Potong data agar sesuai dengan reshape
rss_ip1 = rss_ip1[:aa * bb]
rss_ip2 = rss_ip2[:aa * bb]
rss_ip3 = rss_ip3[:aa * bb]
rss_ip4 = rss_ip4[:aa * bb]

# Melakukan reshape data untuk IP1, IP2, IP3, dan IP4
ip1 = np.array(rss_ip1).reshape(aa, bb).T
ip2 = np.array(rss_ip2).reshape(aa, bb).T
ip3 = np.array(rss_ip3).reshape(aa, bb).T
ip4 = np.array(rss_ip4).reshape(aa, bb).T

# Parameter Kalman Filter
a = 1
h = 1
R = 1
Q = 0.01

xaposteriori_0 = -5
paposteriori_0 = 1

# Kalman Filter
def kalman_filter(rssival):
    xapriori = []
    residual = []
    papriori = []
    k = []
    paposteriori = []
    xaposteriori = []

    row1, row2, row3, row4, row5, row6 = [], [], [], [], [], []
    for m in range(0, aa):
        row1.append(a * xaposteriori_0)
        row2.append(rssival[0][m] - h * row1[m])
        row3.append(a * a * paposteriori_0 + Q)
        row4.append(row3[m] / (row3[m] + R))
        row5.append(row3[m] * (1 - row4[m]))
        row6.append(row1[m] + row4[m] * row2[m])
    xapriori.append(row1)
    residual.append(row2)
    papriori.append(row3)
    k.append(row4)
    paposteriori.append(row5)
    xaposteriori.append(row6)

    for j in range(1, bb):
        row7, row8, row9, row10, row11, row12 = [], [], [], [], [], []
        for n in range(0, aa):
            row7.append(xaposteriori[j - 1][n])
            row8.append(rssival[j][n] - h * row7[n])
            row9.append(a * a * paposteriori[j - 1][n] + Q)
            row10.append(row9[n] / (row9[n] + R))
            row11.append(row9[n] * (1 - row10[n]))
            row12.append(row7[n] + row10[n] * row8[n])
        xapriori.append(row7)
        residual.append(row8)
        papriori.append(row9)
        k.append(row10)
        paposteriori.append(row11)
        xaposteriori.append(row12)
    
    return np.array(xaposteriori)

# Kalman Filter untuk IP1, IP2, IP3, IP4
kalman_ip1 = kalman_filter(ip1).T.flatten()
kalman_ip2 = kalman_filter(ip2).T.flatten()
kalman_ip3 = kalman_filter(ip3).T.flatten()
kalman_ip4 = kalman_filter(ip4).T.flatten()

# Menyimpan hasil ke file Excel menggunakan openpyxl
output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active
output_sheet.title = "KalmanFilter"

# Menulis header
output_sheet.append(['IP1', 'IP2', 'IP3', 'IP4'])

# Menulis data hasil Kalman Filter
for i in range(len(kalman_ip1)):
    output_sheet.append([kalman_ip1[i], kalman_ip2[i], kalman_ip3[i], kalman_ip4[i]])

output_workbook.save('Praproses_Kalman_IP1_IP2_IP3_IP4.xlsx')

end1 = time.time()
time_kalman = end1 - start1
print('Waktu komputasi kalman = {} seconds'.format(time_kalman))

# Visualisasi hasil Kalman Filter
def plot_kalman_results(original, kalman_filtered, label):
    plt.figure(figsize=(10, 6))
    plt.plot(original, label=f'Original {label}', linestyle='--', alpha=0.7)
    plt.plot(kalman_filtered, label=f'Kalman Filtered {label}', linewidth=2)
    plt.title(f'Hasil Kalman Filter - {label}')
    plt.xlabel('Data Index')
    plt.ylabel('Value')
    plt.legend()
    plt.grid(True)
    plt.show()

# Plot untuk IP1
plot_kalman_results(rss_ip1, kalman_ip1, 'IP1')

# Plot untuk IP2
plot_kalman_results(rss_ip2, kalman_ip2, 'IP2')

# Plot untuk IP3
plot_kalman_results(rss_ip3, kalman_ip3, 'IP3')

# Plot untuk IP4
plot_kalman_results(rss_ip4, kalman_ip4, 'IP4')
