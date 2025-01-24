import numpy
import numpy as np
import time
import csv
import matplotlib.pyplot as plt
import pandas as pd
from pandas import read_excel
import math
import subprocess
import openpyxl  # Use openpyxl instead of xlrd
from math import floor
from math import pow
from math import log
from random import seed, randint
from tempfile import TemporaryFile
from sys import exit
from xlwt import Workbook
#import keras.models (Uncomment if using Keras)
#import keras.layers (Uncomment if using Keras)
#import sklearn.preprocessing (Uncomment if using scikit-learn)
import sys
import hashlib
import socket
import pyaes
import binascii

#=====================================================================================================================
#============================================BCH CODE=================================================================
print("\n================================= BCH CODE ===============================\n")
start4 = time.time()

# Replace xlrd with openpyxl for .xlsx files
workbook = openpyxl.load_workbook("E:\\PENS\\Semester 7\\Progress PA\\codingan\\Program\\kuantisasi_doss1_alice.xlsx")
worksheet = workbook.active  # Access the active sheet, assuming it's the first sheet
workbook1 = openpyxl.load_workbook("E:\\PENS\\Semester 7\\Progress PA\\codingan\\Program\\kuantisasi_doss2_bob.xlsx")
worksheet1 = workbook1.active  # Access the active sheet, assuming it's the first sheet

# Read header
first_row = []  # Header
for col in range(worksheet.max_column):
    first_row.append(worksheet.cell(row=1, column=col+1).value)

a = []
for row in range(2, worksheet.max_row + 1):
    elm = worksheet.cell(row=row, column=1).value
    a.append(elm)

b = []
for row in range(2, worksheet1.max_row + 1):
    elm = worksheet1.cell(row=row, column=1).value
    b.append(elm)

alalice = []
bobob = []
deleteblok = 0
m = 5
n = 2**m - 1  # codeword length (n = 2^m - 1 = 31)
k = 6  # information bits
t = 7  # correcting capability

maks = math.floor(len(a)/k)
alice = []
bob = []
for i in range(0, int(maks)):
    alice.append(a[(k*i):(k*(i+1))])
    bob.append(b[(k*i):(k*(i+1))])

# ------------------- GF-ARITHMETIC --------------------
# ------------------------------------------------------
def gf_add(a, b):
    prod = ""
    if len(a) > len(b):
        b = '0'*(len(a) - len(b)) + b
    elif len(a) < len(b):
        a = '0'*(len(b) - len(a)) + a
    for i in range(len(a)):
        if a[i] == b[i]:
            prod += '0'
        else:
            prod += '1'
    return prod


def gf_mul(a, b):
    prod = '0'*(len(a) + len(b) - 1)
    for i in range(len(a)):
        for j in range(len(b)):
            if a[i] == '0' or b[j] == '0':
                tmp = '0'
            else:
                tmp = '1'
            if prod[i+j] == tmp:
                prod = prod[:i+j] + '0' + prod[i+j+1:]
            else:
                prod = prod[:i+j] + '1' + prod[i+j+1:]
    return prod


def gf_xor(a, b):
    prod = ""
    for i in range(1, len(b)):
        if a[i] == b[i]:
            prod += '0'
        else:
            prod += '1'
    return prod


def gf_div(a, b, mode=0):
    pick = len(b)
    q = ""
    r = a[:pick]
    while pick < len(a):
        if r[0] == '1':
            q += '1'
            r = gf_xor(b, r) + a[pick]
        else:
            q += '0'
            r = gf_xor('0'*pick, r) + a[pick]
        pick += 1
    if r[0] == '1':
        q += '1'
        r = gf_xor(b, r)
    else:
        q += '0'
        r = gf_xor('0'*pick, r)
    if not mode:
        return r
    return q


# ------------------- CODER/DECODER --------------------
# ------------------------------------------------------
def bch_encode(data, g):
    return gf_mul(data, g)

def bch_decode(c, g):
    syndrome = gf_div(c, g)
    if not weight(syndrome):
        return c
    cnt_rot = 0
    recd = c
    stp = 0
    while weight(syndrome) > t:
        recd = l_rotate(recd)
        syndrome = gf_div(recd, g)
        cnt_rot += 1
        stp += 1
        if stp > n:
            break
    recd = gf_add(recd, syndrome)
    recd = r_rotate(recd, cnt_rot)
    return recd

# --------------------- UTILITIES ----------------------
# ------------------------------------------------------
def l_rotate(poly, s=1):
    return poly[s:] + poly[:s]

def r_rotate(poly, s=1):
    return poly[-s:] + poly[:-s]

def weight(poly):
    return sum([int(coeff) for coeff in poly])

def polynomial(poly):
    for coeff in range(len(poly)):
        if (poly[coeff] != '0'):
            if (coeff != len(poly)-1):
                print("x^" + str(len(poly)-coeff-1) + " + ", end="")
            else:
                print("1", end="")
    print()


# ------------------------ MAIN ------------------------
# ------------------------------------------------------
global totalerrorbch
totalerrorbch = []
global parityalice
parityalice = []


def main(data, data2, z):
    global deleteblok
    seed()

    g = "11001011011110101000100111"  # BCH (31,6) m=5 k=6 t=7
    c1 = bch_encode(data, g)
    c2 = bch_encode(data2, g)

    errcode = 0
    poserrcode = []
    for i in range(n-k):
        if c1[i+k] != c2[i+k]:
            errcode += 1
            poserrcode.append(i+k)
    totalerrorbch.append(errcode)
    parityalice.append(c1[k:])
    if errcode <= t:
        for i in range(errcode):
            if c2[poserrcode[i]] == '1':
                c2 = c2[:poserrcode[i]] + '0' + c2[poserrcode[i]+1:]
            else:
                c2 = c2[:poserrcode[i]] + '1' + c2[poserrcode[i]+1:]

        recd1 = bch_decode(c1, g)
        recd2 = bch_decode(c2, g)
        recd1 = gf_div(recd1, g, 1)
        recd2 = gf_div(recd2, g, 1)
        alalice.append(recd1)
        bobob.append(recd2)
    else:
        deleteblok += 1


# ----------------------- START ------------------------
# ------------------------------------------------------
for i in range(int(maks)):
    dt1 = ''.join(str(e) for e in alice[i])
    dt2 = ''.join(str(e) for e in bob[i])
    main(dt1, dt2, i+1)

alice11 = ''.join(alalice)
bob11 = ''.join(bobob)
bitalice = list(map(int, alice11))
bitbob = list(map(int, bob11))

salah = 0
for i in range(len(bitbob)):
    if bitalice[i] != bitbob[i]:
        salah += 1

r1 = len(b)
r2 = len(bitbob)
errbchbefore = []
i = 0
while i < len(b):
    if (a[i] == b[i]):
        i = i + 1
    else:
        errbchbefore.append(i + 1)
        i = i + 1

errbch = []
for i in range(len(bitbob)):
    if (bitalice[i] == bitbob[i]):
        i = i + 1
    else:
        errbch.append(i + 1)

kdrbch = len(errbch) / len(bitbob)
print('BCH jumlah blok error = %d, \nJumlah blok dikoreksi = %d, \nKDR = %f' % (deleteblok, (maks - deleteblok), kdrbch))
print('Jumlah delete blok = %d dari %d blok, \nTotal bit = %d sebelumnya %d , \nTotal error = %d\n' % (deleteblok, maks, len(bitbob), maks * k, sum(totalerrorbch)))
print('Bit proses = %d' % (maks * k))
print('Blok = %d' % maks)
print('Total error = %d' % (sum(totalerrorbch)))
print('Blok koreksi = %d' % (maks - deleteblok))
print('Del blok = %d' % deleteblok)
print('JUMLAH BIT HASIL BCH CODE BOB %d' % len(bitbob))

# SAVING BCH
book = Workbook()
sheet1 = book.add_sheet('HasilBCHbob')
sheet1.write(0, 0, 'Bob')
for i in range(1, len(bitalice) + 1):
    sheet1.write(i, 0, float(bitbob[i - 1]))
book.save('Hasil_BCH_Bob_doss2.xls')
book.save(TemporaryFile())
end4 = time.time()

waktu_kuan = 0.3623464107513428
kgrbch = len(bitalice) * ((waktu_kuan) + (end4 - start4))
print('KGR = %f' % kgrbch)
print('Waktu Proses BCH Code DOSS2 (Bob) : ', end4 - start4)
print("BCH CODE Berhasil")
